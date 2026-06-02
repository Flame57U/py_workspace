from openpyxl import load_workbook

src_wb= load_workbook(r'D:\workspace\source.xlsx')
dest_wb= load_workbook(r"D:\workspace\target.xlsx")

src_ws = src_wb.active
dest_ws = dest_wb.active

for i, col_dim in src_ws.column_dimensions.items():
    src_ws.column_dimensions[i].width = col_dim.width

dest_wb.save(r'D:\workspace\target_with_width.xlsx')